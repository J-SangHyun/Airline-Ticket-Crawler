# -*- coding: utf-8 -*-
import tkinter as tk
import tkinter.ttk as ttk
from datetime import date, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from crawler import get_price
from constant import city2code, cityEurope, cityAmerica


if __name__ == '__main__':
    main = tk.Tk()
    main.title('Airline Ticket Crawler')
    main.resizable(False, False)

    day_list = ['일', '월', '화', '수', '목', '금', '토']
    origin_list = [
        ['인천/김포', '서울특별시'],
        ['인천', 'ICN'],
        ['김포', 'GMP']
    ]
    total_airline_list = \
        ['뉴질랜드항공', '대한항공', '델타항공', '방콕항공', '베트남항공', '싱가포르항공', '아시아나', '에어부산', '에어서울',
         '유나이티드항공', '일본항공', '전일본공수', '제주항공', '중화항공', '진에어', '캐세이패시픽항공', '티웨이항공',
         '필리핀항공', 'MIAT 몽골항공', 'THAI']

    origin_checked = tk.IntVar()
    origin_buttons = [tk.Radiobutton(main, text=origin_list[i][0], value=i, variable=origin_checked)
                      for i in range(len(origin_list))]
    for i in range(len(origin_buttons)):
        origin_buttons[i].place(x=10, y=20*i+10)

    airline_checked = [tk.IntVar() for _ in range(len(total_airline_list))]
    airline_buttons = [tk.Checkbutton(main, text=total_airline_list[i], variable=airline_checked[i])
                       for i in range(len(total_airline_list))]
    for i in range(len(airline_buttons)):
        airline_buttons[i].place(x=200, y=20*i+10)

    cities = list(city2code.keys())
    city_checked = [tk.IntVar() for _ in range(len(cities))]
    city_buttons = [tk.Checkbutton(main, text=cities[i], variable=city_checked[i])
                    for i in range(len(cities))]
    for i in range(len(city_buttons)):
        city_buttons[i].place(x=390, y=20*i+10)

    main.geometry(f'830x{20 + max(len(cities), len(total_airline_list)) * 20}')

    n_dates = 8
    dates_from = [tk.Entry(main, width=9) for _ in range(n_dates)]
    dates_from_t = [tk.Label(main, text='부터') for _ in range(n_dates)]
    dates_period = [tk.Entry(main, width=2) for _ in range(n_dates)]
    dates_period_t = [tk.Label(main, text='일간 검색') for _ in range(n_dates)]
    for i in range(n_dates):
        y = (i+1)*30
        dates_from[i].place(x=570, y=y)
        dates_from_t[i].place(x=640, y=y)
        dates_period[i].place(x=690, y=y)
        dates_period_t[i].place(x=710, y=y)

    progress = ttk.Progressbar(main, length=200)
    progress.place(x=570, y=360)

    def crawling():
        CHROME = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        today = date.today()
        workbook = Workbook()
        origin_idx = origin_checked.get()
        file_name = f'{today.strftime("%Y%m%d")}_{origin_list[origin_idx][0]}.xlsx'

        sheet = workbook.active
        sheet.title = today.strftime('%Y%m%d')
        sheet.cell(row=2, column=2, value='출발일')
        sheet.cell(row=2, column=1, value='출발지')
        sheet.cell(row=3, column=1, value=origin_list[origin_idx][0])
        workbook.save(file_name)

        airline_list, city_list = [], []
        for i in range(len(airline_checked)):
            if airline_checked[i].get() == 1:
                airline_list.append(total_airline_list[i])
        for i in range(len(cities)):
            if city_checked[i].get() == 1:
                city_list.append(cities[i])

        for i in range(len(city_list)):
            sheet.cell(column=3 + i * len(airline_list), row=1, value=city_list[i])
            sheet.merge_cells(start_column=3 + i * len(airline_list), start_row=1,
                              end_column=2 + (i + 1) * len(airline_list), end_row=1)
            for j in range(len(airline_list)):
                sheet.cell(column=3 + i * len(airline_list) + j, row=2, value=airline_list[j])
        workbook.save(file_name)

        total_dates = 0
        dates = []
        for i in range(n_dates):
            if dates_from[i].get() != '' and dates_period[i].get() != '':
                date_s = int(dates_from[i].get())
                period = int(dates_period[i].get())
                dates.append([date(date_s//10000, (date_s % 10000)//100, date_s % 100), period])
                total_dates += period

        progress['maximum'] = len(city_list) * total_dates
        for i in range(len(city_list)):
            city = city_list[i]
            cell_row = 3
            for j in range(len(dates)):
                date_start, date_period = dates[j]
                for k in range(date_period):
                    period = 6 if (city in cityEurope or city in cityAmerica) else 3
                    date1 = date_start + timedelta(days=k)
                    date2 = date1 + timedelta(days=period)
                    prices = get_price(CHROME, origin_list[origin_idx][1], city, date1.strftime("%m-%d"),
                                               date2.strftime("%m-%d"), airline_list)
                    sheet.cell(column=2, row=cell_row, value=f'{int(date1.strftime("%Y"))}년 '
                                                             f'{int(date1.strftime("%m"))}월 '
                                                             f'{int(date1.strftime("%d"))}일 '
                                                             f'{day_list[int(date1.strftime("%w"))]}요일')
                    for l in range(len(airline_list)):
                        cell_column = 3 + i * len(airline_list) + l
                        price = prices[l]
                        if price != -1:
                            sheet.cell(column=cell_column, row=cell_row, value=prices[l])
                    workbook.save(f'{today.strftime("%Y%m%d")}.xlsx')
                    cell_row += 1
                    progress['value'] += 1
                    progress.update()
        workbook.save(file_name)
        CHROME.quit()
        main.quit()

    pixel = tk.PhotoImage(width=1, height=1)
    start_btn = tk.Button(main, text='작업 시작', image=pixel, width=192, height=50, compound='c', command=crawling)
    start_btn.place(x=570, y=300)
    main.mainloop()
